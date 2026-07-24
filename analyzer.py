"""
analyzer.py
===========
The local engine. Run this (run.bat does it for you) and it:

  * serves the dashboard at  http://127.0.0.1:5000
  * refreshes the PSX stock universe on startup (so dropdowns are current)
  * answers the dashboard's API calls:
        GET /api/health
        GET /api/stocks?refresh=1          -> current KSE-100/50/30/KMI-30 + all
        GET /api/analyze?symbol=OGDC       -> scrape + score one company, live

Every analyze call scrapes PSX *fresh* (subject to a short session cache so a
double-click doesn't hammer the site). Nothing is precomputed and shipped — the
numbers are pulled at the moment you click Analyze.
"""

from __future__ import annotations
import json
import os
import re
import threading
import time
import webbrowser
from typing import Dict

from flask import Flask, jsonify, request, send_file, Response

try:
    from flask_cors import CORS
except Exception:  # noqa: BLE001 - optional; only needed if opened from file://
    CORS = None

import config
import utils
import psx_data
import scraper
import scorer
import predictor

app = Flask(__name__)
if CORS:
    CORS(app)

HERE = os.path.dirname(os.path.abspath(__file__))
DASHBOARD = os.path.join(HERE, "dashboard.html")

# Tiny in-memory analysis cache: symbol -> (timestamp, payload)
_analysis_cache: Dict[str, tuple] = {}
_cache_lock = threading.Lock()
# v4.0 INSTANT-ANALYSIS ARCHITECTURE
# ----------------------------------
# The boot ranking scan already scrapes + scores + predicts every KSE-100
# company. Those finished payloads are PRIMED into the analysis/prediction
# caches (memory + disk), so a Fundamental or Technical click answers
# INSTANTLY — no re-scrape, no waiting. Disk copies survive relaunches:
# a fresh-cache relaunch (no scan) still serves instantly, and anything
# older than the TTL is served stale WHILE a background refresh replaces it.
_STALE_MAX_H = 168            # disk payloads older than 7 days go stale-dead
_refreshing: set = set()      # single-flight guard for background refreshes


def _pay_path(kind: str, symbol: str) -> str:
    return os.path.join(config.CACHE_DIR, f"{kind}_{symbol.upper()}.json")


def _pay_save(kind: str, symbol: str, payload: Dict) -> None:
    try:
        os.makedirs(config.CACHE_DIR, exist_ok=True)
        with open(_pay_path(kind, symbol), "w", encoding="utf-8") as fh:
            json.dump({"saved_at": time.time(), "payload": payload}, fh)
    except Exception:  # noqa: BLE001
        pass


def _pay_load(kind: str, symbol: str):
    """→ (age_seconds, payload) or None."""
    try:
        path = _pay_path(kind, symbol)
        if not os.path.exists(path):
            return None
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        payload = data.get("payload")
        if not payload:
            return None
        return (time.time() - float(data.get("saved_at", 0)), payload)
    except Exception:  # noqa: BLE001
        return None


def prime_caches(symbol: str, analysis: Dict, prediction: Dict) -> None:
    """Called by the recommendations scan for every company it evaluates —
    the same engines' output is stored so individual clicks are instant."""
    symbol = symbol.strip().upper()
    now = time.time()
    with _cache_lock:
        _analysis_cache[symbol] = (now, analysis)
        _prediction_cache[symbol] = (now, prediction)
    _pay_save("analysis", symbol, analysis)
    _pay_save("prediction", symbol, prediction)


def _background_refresh(symbol: str) -> None:
    """Re-scrape one symbol quietly and replace both cached payloads."""
    with _cache_lock:
        if symbol in _refreshing:
            return
        _refreshing.add(symbol)

    def _run():
        try:
            scraped = scraper.scrape_company(symbol)
            fund = scorer.score_company(scraped)
            pred = predictor.predict(scraped, fund)
            pred["fundamental"] = {
                "score": fund.get("score"), "verdict": fund.get("verdict"),
                "highlights": fund.get("highlights"),
                "concerns": fund.get("concerns"),
            }
            prime_caches(symbol, fund, pred)
        except Exception:  # noqa: BLE001
            pass
        finally:
            with _cache_lock:
                _refreshing.discard(symbol)

    threading.Thread(target=_run, daemon=True,
                     name=f"psx-refresh-{symbol}").start()


def _payload_session(payload) -> str:
    """Which trading session a cached payload describes ('' if unknown)."""
    if not isinstance(payload, dict):
        return ""
    for key in ("as_of", "as_of_close_date"):
        v = payload.get(key)
        if v:
            return str(v)[:10]
    for sub in ("prediction", "scraped", "profile"):
        inner = payload.get(sub)
        if isinstance(inner, dict):
            v = inner.get("as_of") or inner.get("_as_of_target")
            if v:
                return str(v)[:10]
    return ""


def _session_ok(payload) -> bool:
    """
    v4.0 — the as-of guard on every cache path.

    A TTL alone cannot protect the as-of contract. A payload built at 16:00
    yesterday is only "a few hours old" this morning, yet it describes the
    wrong session; and the serve-stale path below deliberately hands back
    payloads up to a week old. Both would silently violate the rule that every
    figure comes from the previous trading day's close. So a cached payload is
    only reusable when the session it was built for is still the session the
    user is entitled to see. Unknown session => treat as unusable rather than
    guess.
    """
    sess = _payload_session(payload)
    return bool(sess) and sess == utils.as_of_iso()


def _cached_payload(kind: str, cache: Dict, symbol: str, ttl: float):
    """Instant-serving cache logic shared by analyze + predict:
    memory fresh → serve; disk fresh → serve; disk stale (≤7d) → serve AND
    refresh in the background; otherwise None (caller does the slow path).

    v4.0 — every branch is additionally gated on the payload describing the
    CURRENT as-of session (see _session_ok). A session mismatch drops straight
    through to the slow path and triggers a background refresh, so the first
    request of a new trading day rebuilds instead of replaying yesterday."""
    with _cache_lock:
        hit = cache.get(symbol)
    if hit and (time.time() - hit[0]) < ttl and _session_ok(hit[1]):
        return hit[1]
    disk = _pay_load(kind, symbol)
    if disk is not None:
        age, payload = disk
        if not _session_ok(payload):
            _background_refresh(symbol)     # yesterday's copy — rebuild now
            return None
        if age < ttl:
            with _cache_lock:
                cache[symbol] = (time.time() - age, payload)
            return payload
        if age < _STALE_MAX_H * 3600:
            with _cache_lock:
                cache[symbol] = (time.time(), payload)   # stop re-reading disk
            _background_refresh(symbol)                  # freshen quietly
            return payload
    return None


# ---------------------------------------------------------------------------
# Static
# ---------------------------------------------------------------------------
@app.route("/")
def index() -> Response:
    if os.path.exists(DASHBOARD):
        return send_file(DASHBOARD)
    return Response("dashboard.html not found next to analyzer.py", status=500)


@app.route("/api/health")
def health():
    """v4.0 — publishes the as-of session so the UI (and any human reading the
    endpoint) can verify at a glance which trading day the data describes."""
    return jsonify({"ok": True, "app": config.APP_NAME,
                    "version": config.APP_VERSION, "time": utils.now_iso(),
                    "psx_time": utils.psx_now().isoformat(),
                    "as_of": utils.as_of_iso(),
                    "as_of_note": ("All figures describe the close of this "
                                   "trading session (the previous trading day).")})


@app.route("/api/progress")
def progress():
    """v3.5 — real pipeline progress for the dashboard's percentage bar."""
    symbol = request.args.get("symbol", "").strip().upper()
    if not symbol:
        return jsonify({"pct": 0, "stage": ""})
    return jsonify(utils.progress_get(symbol))


# ---------------------------------------------------------------------------
# Stock universe
# ---------------------------------------------------------------------------
@app.route("/api/stocks")
def stocks():
    force = request.args.get("refresh") in ("1", "true", "yes")
    uni = psx_data.get_universe(force_refresh=force)
    return jsonify(uni)


# ---------------------------------------------------------------------------
# Analyse one company
# ---------------------------------------------------------------------------
def _analyse(symbol: str) -> Dict:
    symbol = symbol.strip().upper()
    ttl = config.ANALYSIS_TTL_MINUTES * 60
    # v4.0 — INSTANT path: payload primed by the boot scan (memory or disk;
    # a stale-but-recent disk copy is served immediately while a background
    # refresh replaces it)
    cached = _cached_payload("analysis", _analysis_cache, symbol, ttl)
    if cached is not None:
        utils.progress_update(symbol, 100, "Ready")
        return cached

    scraped = scraper.scrape_company(symbol)
    utils.progress_update(symbol, 90, "Scoring the 11 fundamentals…")
    result = scorer.score_company(scraped)
    utils.progress_update(symbol, 100, "Ready")

    with _cache_lock:
        _analysis_cache[symbol] = (time.time(), result)
    _pay_save("analysis", symbol, result)
    return result


@app.route("/api/analyze")
def analyze():
    symbol = request.args.get("symbol", "").strip()
    if not symbol:
        return jsonify({"error": "Pass ?symbol=TICKER"}), 400
    try:                                   # v3.3: background work yields
        import deepdata; deepdata.note_user_activity()
    except Exception:  # noqa: BLE001
        pass
    if request.args.get("fresh") in ("1", "true"):
        with _cache_lock:
            _analysis_cache.pop(symbol.upper(), None)
    try:
        return jsonify(_analyse(symbol))
    except Exception as exc:  # noqa: BLE001
        return jsonify({"error": f"Analysis failed: {exc}", "symbol": symbol}), 500


# ---------------------------------------------------------------------------
# Prediction (technical + fundamental outlook — guidance only, never advice)
# ---------------------------------------------------------------------------
_prediction_cache: Dict[str, tuple] = {}


@app.route("/api/predict")
def predict_route():
    """
    Scrape (or reuse the cached scrape), score fundamentals, then run the
    Bulls-&-Bears-style prediction engine in predictor.py.  The dashboard
    computes the identical analysis client-side from the /api/analyze payload,
    so this endpoint mainly exists for API users and scripting.
    """
    symbol = request.args.get("symbol", "").strip().upper()
    if not symbol:
        return jsonify({"error": "Pass ?symbol=TICKER"}), 400
    try:                                   # v3.3: background work yields
        import deepdata; deepdata.note_user_activity()
    except Exception:  # noqa: BLE001
        pass
    ttl = config.ANALYSIS_TTL_MINUTES * 60
    # v4.0 — INSTANT path (primed by the boot scan; stale disk copies are
    # served immediately while a quiet background refresh runs)
    cached = _cached_payload("prediction", _prediction_cache, symbol, ttl)
    if cached is not None:
        utils.progress_update(symbol, 100, "Ready")
        return jsonify(cached)
    try:
        scraped = scraper.scrape_company(symbol)
        utils.progress_update(symbol, 90, "Running the technical engine…")
        fundamental = scorer.score_company(scraped)
        result = predictor.predict(scraped, fundamental)
        utils.progress_update(symbol, 100, "Ready")
        result["fundamental"] = {
            "score": fundamental.get("score"),
            "verdict": fundamental.get("verdict"),
            "highlights": fundamental.get("highlights"),
            "concerns": fundamental.get("concerns"),
        }
        with _cache_lock:
            _prediction_cache[symbol] = (time.time(), result)
        _pay_save("prediction", symbol, result)
        return jsonify(result)
    except Exception as exc:  # noqa: BLE001
        return jsonify({"error": f"Prediction failed: {exc}", "symbol": symbol}), 500



# ---------------------------------------------------------------------------
# v4.0 — Recommended stocks of the month
# ---------------------------------------------------------------------------
@app.route("/api/recommendations")
def recommendations():
    """Top monthly picks ranked by the blended fundamental + technical score.
    While the background scan is running the payload reports live progress."""
    try:
        import recommend
        force = request.args.get("refresh") in ("1", "true", "yes")
        return jsonify(recommend.get_recommendations(force=force))
    except Exception as exc:  # noqa: BLE001
        return jsonify({"status": "error", "error": str(exc), "picks": []}), 500


# ---------------------------------------------------------------------------
# v4.1 — About this company (sourced, never invented)
# ---------------------------------------------------------------------------
@app.route("/api/about")
def about_route():
    """Full About block for one company. Unlike the copy embedded in an
    analysis payload (which is built from the already-downloaded PSX page so
    the ranking scan stays fast), this endpoint is allowed to consult
    StockAnalysis and the issuer's own website for anything PSX omits."""
    symbol = request.args.get("symbol", "").strip().upper()
    if not symbol:
        return jsonify({"error": "Pass ?symbol=TICKER"}), 400
    try:
        import about
        return jsonify(about.get_about(symbol, allow_fetch=True))
    except Exception as exc:  # noqa: BLE001
        return jsonify({"error": f"About lookup failed: {exc}",
                        "symbol": symbol}), 500


# ---------------------------------------------------------------------------
# v4.1 — Export (PDF / Excel)
# ---------------------------------------------------------------------------
@app.route("/api/export", methods=["POST"])
def export_route():
    """The browser posts the EXACT rows it is showing; this only lays them out.

    Nothing is recomputed server-side, so an export can never disagree with
    the screen. Body: {"format": "pdf"|"xlsx", "rows": [...], "meta": {...}}
    """
    try:
        body = request.get_json(force=True, silent=True) or {}
    except Exception:  # noqa: BLE001
        body = {}
    fmt = (body.get("format") or "").strip().lower()
    rows = body.get("rows") or []
    meta = body.get("meta") or {}

    if fmt not in ("pdf", "xlsx"):
        return jsonify({"error": "format must be 'pdf' or 'xlsx'"}), 400
    if not rows:
        return jsonify({"error": "Nothing to export — no stocks were sent."}), 400

    meta.setdefault("as_of", utils.as_of_iso())
    meta.setdefault("app_version", config.APP_VERSION)

    try:
        import exporter
        data, mime, filename = exporter.build(fmt, rows, meta)
    except RuntimeError as exc:            # a missing optional package
        return jsonify({"error": str(exc)}), 503
    except Exception as exc:  # noqa: BLE001
        return jsonify({"error": f"Export failed: {exc}"}), 500

    resp = Response(data, mimetype=mime)
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.headers["Content-Length"] = str(len(data))
    resp.headers["X-Export-Filename"] = filename
    resp.headers["Access-Control-Expose-Headers"] = "X-Export-Filename"
    return resp


# ---------------------------------------------------------------------------
# v4.1 — Parse an uploaded stock list (CSV / Excel, no header required)
# ---------------------------------------------------------------------------
@app.route("/api/parse-symbols", methods=["POST"])
def parse_symbols_route():
    """The bulk-export panel accepts a plain column of stock names. Anything
    that is not a real PSX symbol is reported back rather than silently
    dropped, so the user can see exactly what was ignored."""
    up = request.files.get("file")
    if up is None or not up.filename:
        return jsonify({"error": "No file was uploaded."}), 400

    name = (up.filename or "").lower()
    raw: list = []
    try:
        if name.endswith((".xlsx", ".xlsm", ".xltx", ".xls")):
            try:
                from openpyxl import load_workbook
            except Exception:  # noqa: BLE001
                return jsonify({"error": "Reading Excel files needs the "
                                         "'openpyxl' package. Save the list as "
                                         "CSV instead, or run: pip install "
                                         "openpyxl"}), 503
            wb = load_workbook(up, read_only=True, data_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            raw.append(str(cell))
        else:
            data = up.read()
            if isinstance(data, bytes):
                for enc in ("utf-8-sig", "utf-8", "latin-1"):
                    try:
                        text = data.decode(enc)
                        break
                    except Exception:  # noqa: BLE001
                        continue
                else:
                    text = data.decode("utf-8", "ignore")
            else:
                text = str(data)
            for line in text.splitlines():
                for cell in line.replace(";", ",").replace("\t", ",").split(","):
                    raw.append(cell)
    except Exception as exc:  # noqa: BLE001
        return jsonify({"error": f"Could not read that file: {exc}"}), 400

    # ---- Match against the real PSX universe -----------------------------
    # A NAME must work exactly as well as a SYMBOL, so names are matched on a
    # normalised form: case, punctuation, "&"/"and" and legal suffixes
    # (Limited / Ltd / Corporation / Pvt …) are all ignored. That way
    # "Lucky Cement Limited", "lucky cement" and "LUCK" all resolve to LUCK.
    uni = psx_data.get_universe()
    resolver = _build_symbol_resolver(uni)

    symbols, unknown, seen = [], [], set()
    for cell in raw:
        token = str(cell).strip().strip('"\'')
        if not token:
            continue
        # a header cell like "Symbol" / "Company" is skipped, not reported
        if token.lower() in ("symbol", "symbols", "stock", "stocks", "ticker",
                             "tickers", "company", "company name", "name",
                             "scrip", "code", "s.no", "sr", "sr.", "#"):
            continue
        hit = resolver(token)
        if hit:
            if hit.upper() not in seen:
                seen.add(hit.upper())
                symbols.append(hit)
        elif token.upper() not in seen:
            seen.add(token.upper())
            unknown.append(token[:40])

    return jsonify({"symbols": symbols, "unknown": unknown[:20],
                    "count": len(symbols)})


# ---------------------------------------------------------------------------
# v4.1 — Symbol-or-name resolution
# ---------------------------------------------------------------------------
# Legal-form words and country tags that appear inconsistently between a PSX
# listing name and whatever a user typed into their spreadsheet.
_NAME_NOISE = re.compile(
    r"\b(limited|ltd|corporation|corp|company|co|incorporated|inc|plc|"
    r"private|pvt|holdings?|group|pakistan|pak)\b", re.I)


def _norm_name(text: str) -> str:
    """Reduce a company name to a comparable core.

    'Lucky Cement Limited', 'LUCKY CEMENT LTD.' and 'lucky  cement' all
    collapse to 'lucky cement', so a name resolves as reliably as a symbol.
    """
    s = (text or "").lower().strip()
    s = s.replace("&", " and ")
    s = re.sub(r"[^a-z0-9 ]+", " ", s)      # drop punctuation
    s = _NAME_NOISE.sub(" ", s)
    return re.sub(r"\s+", " ", s).strip()


def _build_symbol_resolver(uni: Dict):
    """→ resolve(token) -> canonical symbol or None.

    Tries, in order of confidence:
      1. an exact symbol match      (LUCK, luck)
      2. an exact normalised name   ("Lucky Cement Limited" → LUCK)
      3. an UNAMBIGUOUS containment ("lucky cement co" ↔ "lucky cement")
    Step 3 never fires when more than one company could be meant, so a vague
    token is reported back as unknown instead of being silently guessed.
    """
    by_symbol: Dict[str, str] = {}
    by_name: Dict[str, str] = {}
    names: list = []
    for row in (uni.get("symbols") or []):
        sym = (row.get("symbol") or "").strip()
        if not sym:
            continue
        by_symbol[sym.upper()] = sym
        norm = _norm_name(row.get("name") or "")
        if norm:
            # first listing wins so a later duplicate cannot hijack the name
            by_name.setdefault(norm, sym)
            names.append((norm, sym))

    def resolve(token: str):
        token = (token or "").strip()
        if not token:
            return None
        hit = by_symbol.get(token.upper())
        if hit:
            return hit
        norm = _norm_name(token)
        if not norm:
            return None
        hit = by_name.get(norm)
        if hit:
            return hit
        if len(norm) >= 4:
            matches = {s for n, s in names if norm in n or n in norm}
            if len(matches) == 1:
                return matches.pop()
        return None

    return resolve


# ---------------------------------------------------------------------------
# Startup
# ---------------------------------------------------------------------------
def _warm_universe():
    """Refresh the stock list in the background so the first page load is quick."""
    try:
        psx_data.get_universe(force_refresh=True)
    except Exception as exc:  # noqa: BLE001
        print(f"[startup] universe warm-up failed: {exc}")
    # v4.0 speed rule — the landing page's ranked lists come FIRST. Kick the
    # (parallel) recommendations build off before anything else so the main
    # page fills within seconds; a cached list is served instantly meanwhile.
    try:
        import recommend
        recommend.start_build(False)
        # v4.0 backstop — a launch that serves a cached list runs no scan;
        # quietly fill any MISSING analysis payloads so individual clicks
        # are instant even on the first launch after an upgrade.
        recommend.prime_missing()
    except Exception as exc:  # noqa: BLE001
        print(f"[startup] recommendations build not started: {exc}")
    # v3.2 — then the polite background pass that collects official filed
    # reports for every listed COMPANY (ETFs and debt instruments skipped).
    # v4.0.2 speed rule — the prewarm waits until USABLE ranked lists exist
    # (lists_ready), not merely until is_building() flips: checking
    # is_building() had a race — the gate could look BEFORE the build thread
    # started, see False, and release the prewarm to fight the scan for
    # bandwidth. lists_ready() has no such race, and recommend additionally
    # parks the prewarm (note_user_activity) while any scan is running.
    def _prewarm_after_recommend():
        try:
            import recommend
            waited = 0
            while waited < 1800 and not recommend.lists_ready():  # ≤30 min guard
                time.sleep(10)
                waited += 10
        except Exception:  # noqa: BLE001
            pass
        try:
            import deepdata
            uni = psx_data.get_universe()
            syms = [s["symbol"] for s in uni.get("symbols", [])
                    if not s.get("isETF") and not s.get("isDebt")]
            deepdata.start_prewarm(syms)
        except Exception as exc:  # noqa: BLE001
            print(f"[startup] deep-data prewarm not started: {exc}")

    threading.Thread(target=_prewarm_after_recommend, daemon=True,
                     name="psx-prewarm-gate").start()


@app.route("/api/deep-status")
def deep_status():
    """Progress of the official-filings store (deepdata.py)."""
    try:
        import deepdata
        return jsonify(deepdata.status())
    except Exception as exc:  # noqa: BLE001
        return jsonify({"error": str(exc)}), 500


def _open_browser():
    time.sleep(1.2)
    try:
        webbrowser.open(f"http://{config.HOST}:{config.PORT}")
    except Exception:  # noqa: BLE001
        pass


def main():
    print("=" * 60)
    print(f"  {config.APP_NAME}  ·  v{config.APP_VERSION}")
    print(f"  Dashboard:  http://{config.HOST}:{config.PORT}")
    print("  Refreshing PSX stock list in the background ...")
    print("=" * 60)

    threading.Thread(target=_warm_universe, daemon=True).start()
    if config.OPEN_BROWSER:
        threading.Thread(target=_open_browser, daemon=True).start()

    # threaded so background scrapes don't block the UI
    app.run(host=config.HOST, port=config.PORT, threaded=True, debug=False)


if __name__ == "__main__":
    main()
