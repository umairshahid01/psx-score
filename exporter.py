"""
exporter.py
===========
v4.1 — the export engine.

Turns the exact payload the dashboard is *showing* into either

    * an XLSX workbook with two sheets — "Fundamental" and "Technical", or
    * a comprehensive PDF report mirroring both tabs of the tool.

Design rules
------------
1.  NOTHING IS RECOMPUTED HERE. The browser sends the very numbers it rendered,
    so an export can never disagree with the screen. This module only lays them
    out. Missing values stay missing — they are printed as "—", never filled in.

2.  The thirteen MUST-HAVE columns lead BOTH sheets, in the requested order, so
    each sheet is independently useful:

        Stock Symbol · Company Name · Stock Category (PSX Sector) ·
        The energy meter (RSI) · How loud is the crowd? (volume) ·
        Chart Health · Business Health · OverAll Outlook ·
        The market's season (Wyckoff) · Buy zone 1 (first nibble) ·
        Buy zone 2 (deeper dip) · Stop-loss (the safety rope) ·
        Target 2 (if Target 1 breaks)

3.  PDF text is ASCII-safe. The standard PDF fonts have no emoji and no ₨
    glyph, so the report uses "Rs" and colour-codes the trade-plan rows
    (green = buy, red = stop, blue = target) instead of coloured squares.
    The XLSX keeps the emoji headings exactly as specified.
"""

from __future__ import annotations

import datetime as _dt
import io
import re
from typing import Any, Dict, List, Optional

# ---------------------------------------------------------------------------
# Optional dependencies — the app must still boot if either is missing.
# ---------------------------------------------------------------------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
    from openpyxl.utils import get_column_letter
    HAVE_XLSX = True
except Exception:  # noqa: BLE001
    HAVE_XLSX = False

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (BaseDocTemplate, Frame, KeepTogether,
                                    NextPageTemplate, PageBreak, PageTemplate,
                                    Paragraph, Spacer, Table, TableStyle)
    HAVE_PDF = True
except Exception:  # noqa: BLE001
    HAVE_PDF = False


# ---------------------------------------------------------------------------
# Brand palette (mirrors the dashboard)
# ---------------------------------------------------------------------------
C_INK = "1B2430"
C_MUTED = "6B7A8D"
C_LINE = "D8E0E8"
C_BAND = "F4F7FA"
C_HEAD = "0E3A5C"      # deep PSX blue
C_ACCENT = "1E7F5C"    # PSX green
C_GOOD = "1B7F4B"
C_WARN = "9A6A00"
C_BAD = "B3261E"
C_BUY = "1B7F4B"
C_STOP = "B3261E"
C_TARGET = "1B5FA8"

# ---------------------------------------------------------------------------
# The thirteen must-have columns
# ---------------------------------------------------------------------------
MUST_HAVE: List[tuple] = [
    ("symbol",          "Stock Symbol",                      10),
    ("name",            "Company Name",                      30),
    ("category",        "Stock Category (PSX Sector)",       26),
    ("rsi",             "The energy meter (RSI)",            18),
    ("volume",          "How loud is the crowd? (volume)",   22),
    ("chart_health",    "Chart Health",                      13),
    ("business_health", "Business Health",                   15),
    ("outlook",         "OverAll Outlook",                   24),
    ("wyckoff",         "The market\u2019s season (Wyckoff)", 24),
    ("buy1",            "\U0001F7E2 Buy zone 1 (first nibble)",       20),
    ("buy2",            "\U0001F7E2 Buy zone 2 (deeper dip)",         20),
    ("stop",            "\U0001F6D1 Stop-loss (the safety rope)",     22),
    ("t2",              "\U0001F3AF Target 2 (if Target 1 breaks)",   22),
]

# Plain (emoji-free) headings for the PDF
_PDF_HEAD = {
    "buy1": "Buy zone 1\n(first nibble)",
    "buy2": "Buy zone 2\n(deeper dip)",
    "stop": "Stop-loss\n(the safety rope)",
    "t2": "Target 2\n(if T1 breaks)",
    "rsi": "Energy meter\n(RSI)",
    "volume": "Crowd noise\n(volume)",
    "wyckoff": "Market season\n(Wyckoff)",
    "category": "Category\n(PSX Sector)",
    "chart_health": "Chart\nHealth",
    "business_health": "Business\nHealth",
    "outlook": "Overall\nOutlook",
    "symbol": "Symbol",
    "name": "Company Name",
}

# Sheet-specific detail columns appended after the must-haves
FUND_EXTRA: List[tuple] = [
    ("fund_score",     "Fundamental Score / 100", 20),
    ("fund_verdict",   "Fundamental Verdict",     20),
    ("model",          "Scoring Model",           16),
    ("confidence",     "Data Confidence",         16),
    ("price",          "Share Price (Rs)",        16),
    ("change_pct",     "Day Change %",            13),
    ("market_cap",     "Market Cap (Rs)",         20),
    ("pe",             "P/E Ratio",               11),
    ("week52_low",     "52-week Low (Rs)",        16),
    ("week52_high",    "52-week High (Rs)",       16),
    ("highlights",     "Strengths",               46),
    ("concerns",       "Concerns",                46),
    ("material_info",  "Material Information (latest filings)", 52),
    ("about",          "About the Business",      64),
    ("products",       "Top Products",            36),
    ("website",        "Official Website",        28),
    ("psx_source",     "PSX Source",              34),
]

TECH_EXTRA: List[tuple] = [
    ("structure",   "Trend Structure",            18),
    ("t1",          "\U0001F3AF Target 1",        14),
    ("rr",          "Risk : Reward",              14),
    ("avg_buy",     "Average Buy (Rs)",           16),
    ("entry_state", "Entry Timing",               16),
    ("entry_note",  "Entry Note",                 52),
    ("atr_pct",     "Volatility (ATR %)",         16),
    ("week52_low",  "52-week Low (Rs)",           16),
    ("week52_high", "52-week High (Rs)",          16),
    ("supports",    "Nearest Floors (support)",   26),
    ("resistances", "Nearest Ceilings (resistance)", 26),
    ("divergence",  "RSI Divergence",             18),
    ("pros",        "What's going right",         52),
    ("cons",        "What to watch",              52),
    ("psx_source",  "PSX Source",                 34),
]


# ---------------------------------------------------------------------------
# Value helpers — a missing figure stays missing
# ---------------------------------------------------------------------------
DASH = "\u2014"


def _s(val: Any) -> str:
    """Render any value as display text without inventing anything."""
    if val is None:
        return DASH
    if isinstance(val, float):
        if val != val:                       # NaN
            return DASH
    if isinstance(val, (list, tuple)):
        items = [_s(v) for v in val if v not in (None, "")]
        items = [i for i in items if i != DASH]
        return " \u2022 ".join(items) if items else DASH
    text = str(val).strip()
    return text if text else DASH


def _ascii(text: Any) -> str:
    """PDF-safe text: strip emoji/symbols the base-14 fonts cannot draw."""
    t = _s(text)
    t = t.replace("\u20a8", "Rs ").replace("\u20b9", "Rs ")
    t = (t.replace("\u2014", "-").replace("\u2013", "-")
           .replace("\u2018", "'").replace("\u2019", "'")
           .replace("\u201c", '"').replace("\u201d", '"')
           .replace("\u2022", "*").replace("\u2026", "..."))
    # drop anything outside Latin-1 (emoji, arrows, CJK…)
    t = "".join(ch if ord(ch) < 256 else " " for ch in t)
    return re.sub(r"[ \t]{2,}", " ", t).strip()


def _num(val: Any) -> Optional[float]:
    try:
        if val is None or val == "":
            return None
        f = float(val)
        return None if f != f else f
    except Exception:  # noqa: BLE001
        return None


def _money(val: Any, prefix: str = "\u20a8 ") -> str:
    """Share prices always carry 2 decimals so a buy/stop/target column lines
    up visually — ₨ 980.00 next to ₨ 1,240.00, never ₨ 1,240."""
    n = _num(val)
    if n is None:
        return DASH
    return f"{prefix}{n:,.2f}"


def _money_ascii(val: Any) -> str:
    n = _num(val)
    return DASH if n is None else f"Rs {n:,.2f}"


def _big_money(val: Any) -> str:
    n = _num(val)
    if n is None:
        return DASH
    for cut, suf in ((1e12, "trn"), (1e9, "bn"), (1e6, "mn")):
        if abs(n) >= cut:
            return f"\u20a8 {n / cut:,.2f} {suf}"
    return f"\u20a8 {n:,.0f}"


def _score(val: Any) -> str:
    n = _num(val)
    return DASH if n is None else f"{n:.0f} / 100"


# ---------------------------------------------------------------------------
# Row flattening — one dashboard row -> the flat cells each sheet needs
# ---------------------------------------------------------------------------
def _flat(row: Dict, ascii_only: bool = False) -> Dict[str, str]:
    """Build every cell value once; both sheets and the PDF read from here."""
    row = row or {}
    f = row.get("fund") or {}
    t = row.get("tech") or {}
    a = row.get("about") or {}
    M = _money_ascii if ascii_only else _money

    rsi = _num(t.get("rsi"))
    rsi_txt = DASH if rsi is None else f"{rsi:.0f}"
    if rsi is not None and t.get("rsi_tag"):
        rsi_txt = f"{rsi:.0f} \u2014 {t.get('rsi_tag')}"

    vol = t.get("volume_note")
    vol_txt = _s(vol)
    if vol_txt != DASH:
        vol_txt = vol_txt[:1].upper() + vol_txt[1:]

    wy = t.get("wyckoff_label") or t.get("wyckoff_phase")

    out = {
        # --- must-haves ---------------------------------------------------
        "symbol": _s(row.get("symbol")),
        "name": _s(row.get("name")),
        "category": _s(row.get("category")),
        "rsi": rsi_txt,
        "volume": vol_txt,
        "chart_health": _score(t.get("chart_health")),
        "business_health": _score(t.get("business_health") if t.get("business_health") is not None
                                  else f.get("score")),
        "outlook": _s(t.get("outlook_label")),
        "wyckoff": _s(wy),
        "buy1": M(t.get("buy1")),
        "buy2": M(t.get("buy2")),
        "stop": M(t.get("stop")),
        "t2": M(t.get("t2")),
        # --- fundamental detail ------------------------------------------
        "fund_score": _score(f.get("score")),
        "fund_verdict": _s(f.get("verdict_label")),
        "model": _s(f.get("model")),
        "confidence": (DASH if _num(f.get("confidence")) is None
                       else f"{_num(f.get('confidence')):.0f}%"),
        "price": M(f.get("price")),
        "change_pct": (DASH if _num(f.get("change_pct")) is None
                       else f"{_num(f.get('change_pct')):+.2f}%"),
        "market_cap": (_money_ascii(f.get("market_cap")) if ascii_only
                       else _big_money(f.get("market_cap"))),
        "pe": (DASH if _num(f.get("pe")) is None else f"{_num(f.get('pe')):.2f}x"),
        "highlights": _s(f.get("highlights")),
        "concerns": _s(f.get("concerns")),
        "material_info": _s(f.get("material_info_summary")),
        "about": _s(a.get("description")),
        "products": _s(a.get("products")),
        "website": _s(a.get("website")),
        # --- technical detail --------------------------------------------
        "structure": _s(t.get("structure")),
        "t1": M(t.get("t1")),
        "rr": (DASH if _num(t.get("rr")) is None else f"{_num(t.get('rr')):.2f} : 1"),
        "avg_buy": M(t.get("avg_buy")),
        "entry_state": _s(t.get("entry_state")),
        "entry_note": _s(t.get("entry_note")),
        "atr_pct": (DASH if _num(t.get("atr_pct")) is None
                    else f"{_num(t.get('atr_pct')):.2f}%"),
        "week52_low": M(t.get("week52_low") if t.get("week52_low") is not None
                        else f.get("week52_low")),
        "week52_high": M(t.get("week52_high") if t.get("week52_high") is not None
                         else f.get("week52_high")),
        "supports": _s(t.get("supports")),
        "resistances": _s(t.get("resistances")),
        "divergence": _s(t.get("divergence")),
        "pros": _s(t.get("pros")),
        "cons": _s(t.get("cons")),
        "psx_source": _s(row.get("psx_url")),
    }
    if ascii_only:
        out = {k: _ascii(v) for k, v in out.items()}
    return out


def _status_of(row: Dict, key: str) -> str:
    """good / warn / bad for colour-coding a must-have cell."""
    t = (row or {}).get("tech") or {}
    f = (row or {}).get("fund") or {}
    if key == "chart_health":
        n = _num(t.get("chart_health"))
    elif key == "business_health":
        n = _num(t.get("business_health"))
        if n is None:
            n = _num(f.get("score"))
    elif key == "rsi":
        n = _num(t.get("rsi"))
        if n is None:
            return ""
        return "warn" if (n >= 70 or n <= 35) else "good"
    elif key == "outlook":
        n = _num(t.get("combined"))
    else:
        return ""
    if n is None:
        return ""
    return "good" if n >= 60 else ("warn" if n >= 45 else "bad")


# ===========================================================================
#  EXCEL
# ===========================================================================
def _xl_style_header(ws, headers: List[tuple], row_idx: int = 1) -> None:
    fill = PatternFill("solid", fgColor=C_HEAD)
    font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=C_LINE)
    for c, (_key, label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=c, value=label)
        cell.fill, cell.font, cell.alignment = fill, font, align
        cell.border = Border(bottom=Side(style="medium", color=C_ACCENT),
                             left=thin, right=thin)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[row_idx].height = 42


def _xl_write_sheet(ws, title_text: str, subtitle: str,
                    headers: List[tuple], rows: List[Dict]) -> None:
    ncols = len(headers)

    # ---- title band ------------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.font = Font(bold=True, size=16, color=C_HEAD, name="Calibri")
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sc = ws.cell(row=2, column=1, value=subtitle)
    sc.font = Font(size=9, italic=True, color=C_MUTED, name="Calibri")
    sc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    head_row = 4
    _xl_style_header(ws, headers, head_row)

    thin = Side(style="thin", color=C_LINE)
    band = PatternFill("solid", fgColor=C_BAND)
    base_font = Font(size=10, name="Calibri", color=C_INK)
    wrap_keys = {"highlights", "concerns", "about", "products", "entry_note",
                 "pros", "cons", "material_info", "supports", "resistances",
                 "psx_source", "website"}

    for r, row in enumerate(rows, start=head_row + 1):
        flat = _flat(row)
        striped = ((r - head_row) % 2 == 0)
        for c, (key, _label, _w) in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=flat.get(key, DASH))
            cell.font = base_font
            cell.border = Border(bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(
                horizontal="left" if key in wrap_keys or key in ("name", "category")
                else "center",
                vertical="top" if key in wrap_keys else "center",
                wrap_text=key in wrap_keys)
            if striped:
                cell.fill = band
            # identity + verdict emphasis
            if key == "symbol":
                cell.font = Font(bold=True, size=10, name="Calibri", color=C_HEAD)
            elif key in ("buy1", "buy2"):
                cell.font = Font(size=10, name="Calibri", color=C_BUY, bold=True)
            elif key == "stop":
                cell.font = Font(size=10, name="Calibri", color=C_STOP, bold=True)
            elif key in ("t1", "t2"):
                cell.font = Font(size=10, name="Calibri", color=C_TARGET, bold=True)
            else:
                st = _status_of(row, key)
                if st:
                    col = {"good": C_GOOD, "warn": C_WARN, "bad": C_BAD}[st]
                    cell.font = Font(size=10, name="Calibri", color=col, bold=True)
            if key == "psx_source" and flat.get(key) not in (DASH, ""):
                cell.hyperlink = flat[key]
                cell.font = Font(size=9, name="Calibri", color="1155CC",
                                 underline="single")
            if key == "website" and flat.get(key) not in (DASH, ""):
                cell.hyperlink = flat[key]
                cell.font = Font(size=9, name="Calibri", color="1155CC",
                                 underline="single")
        ws.row_dimensions[r].height = 30 if any(
            k in wrap_keys for k, _l, _w in headers) else 18

    # ---- usability -------------------------------------------------------
    ws.freeze_panes = ws.cell(row=head_row + 1, column=4)
    if rows:
        ws.auto_filter.ref = (f"A{head_row}:"
                              f"{get_column_letter(ncols)}{head_row + len(rows)}")
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = f"{head_row}:{head_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_workbook(rows: List[Dict], meta: Dict = None) -> bytes:
    """Two-sheet XLSX: Fundamental + Technical."""
    if not HAVE_XLSX:
        raise RuntimeError(
            "Excel export needs the 'openpyxl' package. Close the app, run:\n"
            "    pip install openpyxl\nand launch again.")
    meta = meta or {}
    stamp = meta.get("as_of") or _dt.date.today().isoformat()
    mode = meta.get("mode") or ""
    sub = (f"Figures describe the close of {stamp}"
           + (f"  \u00b7  {mode} data" if mode else "")
           + f"  \u00b7  {len(rows)} compan{'y' if len(rows) == 1 else 'ies'}"
           + "  \u00b7  exported from PSX\u00b7SCORE. Nothing here is estimated \u2014 "
             "every figure is the one shown in the tool.")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Fundamental"
    _xl_write_sheet(ws1, "PSX\u00b7SCORE \u2014 Fundamental Analysis", sub,
                    MUST_HAVE + FUND_EXTRA, rows)

    ws2 = wb.create_sheet("Technical")
    _xl_write_sheet(ws2, "PSX\u00b7SCORE \u2014 Technical Analysis", sub,
                    MUST_HAVE + TECH_EXTRA, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
#  PDF
# ===========================================================================
def _pdf_styles():
    ss = getSampleStyleSheet()
    return {
        "h1": ParagraphStyle("h1", parent=ss["Title"], fontName="Helvetica-Bold",
                             fontSize=20, leading=24,
                             textColor=colors.HexColor("#" + C_HEAD),
                             alignment=TA_LEFT, spaceAfter=2),
        "sub": ParagraphStyle("sub", parent=ss["Normal"], fontName="Helvetica",
                              fontSize=8.5, leading=12,
                              textColor=colors.HexColor("#" + C_MUTED)),
        "h2": ParagraphStyle("h2", parent=ss["Heading2"],
                             fontName="Helvetica-Bold", fontSize=13, leading=16,
                             textColor=colors.HexColor("#" + C_HEAD),
                             spaceBefore=10, spaceAfter=4),
        "h3": ParagraphStyle("h3", parent=ss["Heading3"],
                             fontName="Helvetica-Bold", fontSize=10, leading=13,
                             textColor=colors.HexColor("#" + C_ACCENT),
                             spaceBefore=8, spaceAfter=3),
        "body": ParagraphStyle("body", parent=ss["Normal"], fontName="Helvetica",
                               fontSize=8.8, leading=12.4,
                               textColor=colors.HexColor("#" + C_INK)),
        "cell": ParagraphStyle("cell", parent=ss["Normal"], fontName="Helvetica",
                               fontSize=7.4, leading=9.4,
                               textColor=colors.HexColor("#" + C_INK)),
        "cellb": ParagraphStyle("cellb", parent=ss["Normal"],
                                fontName="Helvetica-Bold", fontSize=7.4,
                                leading=9.4,
                                textColor=colors.HexColor("#" + C_INK)),
        "th": ParagraphStyle("th", parent=ss["Normal"],
                             fontName="Helvetica-Bold", fontSize=7.2,
                             leading=8.8, textColor=colors.white,
                             alignment=TA_CENTER),
        "note": ParagraphStyle("note", parent=ss["Normal"], fontName="Helvetica-Oblique",
                               fontSize=7.6, leading=10,
                               textColor=colors.HexColor("#" + C_MUTED)),
    }


class _Doc(BaseDocTemplate):
    """Landscape summary pages + portrait detail pages, with a running footer."""

    def __init__(self, buf, meta):
        self.meta = meta or {}
        super().__init__(buf, pagesize=landscape(A4),
                         leftMargin=12 * mm, rightMargin=12 * mm,
                         topMargin=12 * mm, bottomMargin=14 * mm,
                         title="PSX-SCORE Report", author="PSX.SCORE")
        lw, lh = landscape(A4)
        pw, ph = A4
        land = PageTemplate(
            id="land",
            frames=[Frame(12 * mm, 14 * mm, lw - 24 * mm, lh - 26 * mm,
                          id="lf", showBoundary=0)],
            onPage=self._decorate, pagesize=landscape(A4))
        port = PageTemplate(
            id="port",
            frames=[Frame(14 * mm, 14 * mm, pw - 28 * mm, ph - 26 * mm,
                          id="pf", showBoundary=0)],
            onPage=self._decorate, pagesize=A4)
        self.addPageTemplates([land, port])

    def _decorate(self, canv, doc):
        w, h = canv._pagesize
        canv.saveState()
        # top hairline
        canv.setStrokeColor(colors.HexColor("#" + C_ACCENT))
        canv.setLineWidth(1.6)
        canv.line(12 * mm, h - 9 * mm, w - 12 * mm, h - 9 * mm)
        # footer
        canv.setFont("Helvetica", 7)
        canv.setFillColor(colors.HexColor("#" + C_MUTED))
        canv.drawString(12 * mm, 8 * mm,
                        _ascii("PSX.SCORE  -  " + str(self.meta.get("app_version", ""))
                               + "  -  data as of " + str(self.meta.get("as_of", ""))))
        canv.drawCentredString(w / 2.0, 8 * mm,
                               "Guidance only - not investment advice.")
        canv.drawRightString(w - 12 * mm, 8 * mm, "Page %d" % canv.getPageNumber())
        canv.restoreState()


def _kv_table(pairs, styles, col_w, ncols=2):
    """A tidy label/value grid."""
    cells = []
    for label, value in pairs:
        cells.append(Paragraph(f"<font color='#{C_MUTED}'>{_ascii(label)}</font>",
                               styles["cell"]))
        cells.append(Paragraph(f"<b>{_ascii(value)}</b>", styles["cell"]))
    rows, widths = [], []
    per = ncols * 2
    for i in range(0, len(cells), per):
        chunk = cells[i:i + per]
        while len(chunk) < per:
            chunk.append("")
        rows.append(chunk)
    unit = col_w / ncols
    for _ in range(ncols):
        widths.extend([unit * 0.42, unit * 0.58])
    t = Table(rows, colWidths=widths, hAlign="LEFT")
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, colors.HexColor("#" + C_LINE)),
    ]))
    return t


def _summary_table(rows, styles, avail_w):
    keys = [k for k, _l, _w in MUST_HAVE]
    weights = {"symbol": 0.62, "name": 1.55, "category": 1.25, "rsi": 0.85,
               "volume": 0.90, "chart_health": 0.66, "business_health": 0.70,
               "outlook": 1.25, "wyckoff": 1.10, "buy1": 0.82, "buy2": 0.82,
               "stop": 0.82, "t2": 0.82}
    total = sum(weights[k] for k in keys)
    widths = [avail_w * weights[k] / total for k in keys]

    head = [Paragraph(_ascii(_PDF_HEAD.get(k, lbl)).replace("\n", "<br/>"),
                      styles["th"]) for k, lbl, _w in MUST_HAVE]
    body = [head]
    cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + C_HEAD)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (3, 1), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#" + C_LINE)),
        ("LINEBELOW", (0, 0), (-1, 0), 1.1, colors.HexColor("#" + C_ACCENT)),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]
    for i, row in enumerate(rows, start=1):
        flat = _flat(row, ascii_only=True)
        line = []
        for k in keys:
            st = _s(flat.get(k))
            if k == "symbol":
                line.append(Paragraph(f"<b>{st}</b>", styles["cellb"]))
            elif k in ("buy1", "buy2"):
                line.append(Paragraph(f"<font color='#{C_BUY}'><b>{st}</b></font>",
                                      styles["cell"]))
            elif k == "stop":
                line.append(Paragraph(f"<font color='#{C_STOP}'><b>{st}</b></font>",
                                      styles["cell"]))
            elif k == "t2":
                line.append(Paragraph(f"<font color='#{C_TARGET}'><b>{st}</b></font>",
                                      styles["cell"]))
            else:
                stt = _status_of(row, k)
                if stt:
                    col = {"good": C_GOOD, "warn": C_WARN, "bad": C_BAD}[stt]
                    line.append(Paragraph(f"<font color='#{col}'><b>{st}</b></font>",
                                          styles["cell"]))
                else:
                    line.append(Paragraph(st, styles["cell"]))
        body.append(line)
        if i % 2 == 0:
            cmds.append(("BACKGROUND", (0, i), (-1, i),
                         colors.HexColor("#" + C_BAND)))
    t = Table(body, colWidths=widths, repeatRows=1, hAlign="LEFT")
    t.setStyle(TableStyle(cmds))
    return t


def _metric_table(metrics, styles, avail_w):
    head = [Paragraph("Metric", styles["th"]), Paragraph("Value", styles["th"]),
            Paragraph("Reading", styles["th"]),
            Paragraph("How it was worked out", styles["th"]),
            Paragraph("Source", styles["th"])]
    body = [head]
    cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + C_HEAD)),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#" + C_LINE)),
        ("ALIGN", (1, 1), (2, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    LBL = {"good": "Good", "warn": "Watch", "bad": "Weak"}
    for i, m in enumerate(metrics or [], start=1):
        st = (m.get("status") or "").lower()
        col = {"good": C_GOOD, "warn": C_WARN, "bad": C_BAD}.get(st, C_MUTED)
        body.append([
            Paragraph(_ascii(m.get("label")), styles["cell"]),
            Paragraph(f"<b>{_ascii(m.get('display'))}</b>", styles["cell"]),
            Paragraph(f"<font color='#{col}'><b>{LBL.get(st, DASH)}</b></font>",
                      styles["cell"]),
            Paragraph(_ascii(m.get("note")), styles["cell"]),
            Paragraph(_ascii(m.get("source_doc")), styles["cell"]),
        ])
        if i % 2 == 0:
            cmds.append(("BACKGROUND", (0, i), (-1, i),
                         colors.HexColor("#" + C_BAND)))
    if len(body) == 1:
        return Paragraph("No fundamental metrics were available for this "
                         "company.", styles["note"])
    w = [avail_w * x for x in (0.22, 0.13, 0.09, 0.36, 0.20)]
    t = Table(body, colWidths=w, repeatRows=1, hAlign="LEFT")
    t.setStyle(TableStyle(cmds))
    return t


def _plan_table(flat, styles, avail_w):
    rows = [
        ("Buy zone 1 (first nibble)", flat["buy1"], C_BUY),
        ("Buy zone 2 (deeper dip)", flat["buy2"], C_BUY),
        ("Average buy", flat["avg_buy"], C_INK),
        ("Stop-loss (the safety rope)", flat["stop"], C_STOP),
        ("Target 1", flat["t1"], C_TARGET),
        ("Target 2 (if Target 1 breaks)", flat["t2"], C_TARGET),
        ("Risk : Reward", flat["rr"], C_INK),
        ("Entry timing", flat["entry_state"], C_INK),
    ]
    body, cmds = [], [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#" + C_LINE)),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
    ]
    for i, (label, val, col) in enumerate(rows):
        body.append([Paragraph(_ascii(label), styles["cell"]),
                     Paragraph(f"<font color='#{col}'><b>{_ascii(val)}</b></font>",
                               styles["cell"])])
        if i % 2 == 1:
            cmds.append(("BACKGROUND", (0, i), (-1, i),
                         colors.HexColor("#" + C_BAND)))
    t = Table(body, colWidths=[avail_w * 0.55, avail_w * 0.45], hAlign="LEFT")
    t.setStyle(TableStyle(cmds))
    return t


def _bullets(items, styles, empty="Nothing noted."):
    if isinstance(items, str):
        items = [x.strip() for x in re.split(r"\s*\u2022\s*|\n", items) if x.strip()]
    items = [i for i in (items or []) if _s(i) != DASH]
    if not items:
        return Paragraph(empty, styles["note"])
    html = "<br/>".join("&bull; " + _ascii(i) for i in items[:14])
    return Paragraph(html, styles["body"])


def _stock_section(row, styles, avail_w):
    flat = _flat(row, ascii_only=True)
    f = row.get("fund") or {}
    t = row.get("tech") or {}
    a = row.get("about") or {}
    story = []

    # ---- name band -------------------------------------------------------
    band = Table([[Paragraph(
        f"<font size=15 color='white'><b>{_ascii(row.get('symbol'))}</b></font>"
        f"&nbsp;&nbsp;<font size=11 color='white'>{_ascii(row.get('name'))}</font>"
        f"<br/><font size=7.5 color='#CFE3F2'>{_ascii(row.get('category'))}</font>",
        styles["body"])]], colWidths=[avail_w])
    band.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#" + C_HEAD)),
        ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story += [band, Spacer(1, 7)]

    # ---- about -----------------------------------------------------------
    story.append(Paragraph("About the business", styles["h3"]))
    desc = _ascii(a.get("description"))
    story.append(Paragraph(desc if desc != DASH else
                           "No business description is published by PSX, "
                           "StockAnalysis or the company's own website, so none "
                           "is shown here.", styles["body"]))
    prods = a.get("products")
    if prods and _s(prods) != DASH:
        story += [Spacer(1, 3),
                  Paragraph("<b>Top products / lines of business:</b> "
                            + _ascii(prods), styles["body"])]
    if _s(a.get("website")) != DASH:
        story.append(Paragraph("Website: " + _ascii(a.get("website")),
                               styles["note"]))
    story.append(Spacer(1, 5))

    # ---- snapshot --------------------------------------------------------
    story.append(Paragraph("Key snapshot", styles["h3"]))
    story.append(_kv_table([
        ("Energy meter (RSI)", flat["rsi"]),
        ("Crowd noise (volume)", flat["volume"]),
        ("Chart Health", flat["chart_health"]),
        ("Business Health", flat["business_health"]),
        ("Overall Outlook", flat["outlook"]),
        ("Market season (Wyckoff)", flat["wyckoff"]),
        ("Share price", flat["price"]),
        ("Day change", flat["change_pct"]),
    ], styles, avail_w, ncols=2))
    story.append(Spacer(1, 6))

    # ---- fundamental -----------------------------------------------------
    story.append(Paragraph("Fundamental Analysis", styles["h2"]))
    story.append(_kv_table([
        ("Fundamental score", flat["fund_score"]),
        ("Verdict", flat["fund_verdict"]),
        ("Scoring model", flat["model"]),
        ("Data confidence", flat["confidence"]),
        ("Market capitalisation", flat["market_cap"]),
        ("P/E ratio", flat["pe"]),
        ("52-week low", flat["week52_low"]),
        ("52-week high", flat["week52_high"]),
    ], styles, avail_w, ncols=2))
    story.append(Spacer(1, 5))
    story.append(_metric_table(f.get("metrics"), styles, avail_w))

    if f.get("highlights") or f.get("concerns"):
        story.append(Spacer(1, 5))
        left = [Paragraph("Strengths", styles["h3"]),
                _bullets(f.get("highlights"), styles, "None stood out.")]
        right = [Paragraph("Concerns", styles["h3"]),
                 _bullets(f.get("concerns"), styles, "None flagged.")]
        sc = Table([[left, right]], colWidths=[avail_w / 2, avail_w / 2])
        sc.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                                ("RIGHTPADDING", (0, 0), (0, 0), 10)]))
        story.append(sc)

    mi = f.get("material_info_summary")
    if mi and _s(mi) != DASH:
        story += [Spacer(1, 5),
                  Paragraph("Material Information filings", styles["h3"]),
                  _bullets(mi, styles, "No Material Information filings found.")]

    # ---- technical -------------------------------------------------------
    story.append(Paragraph("Technical Analysis", styles["h2"]))
    story.append(_kv_table([
        ("Chart Health", flat["chart_health"]),
        ("Business Health", flat["business_health"]),
        ("Overall Outlook", flat["outlook"]),
        ("Trend structure", flat["structure"]),
        ("Market season (Wyckoff)", flat["wyckoff"]),
        ("Energy meter (RSI)", flat["rsi"]),
        ("Crowd noise (volume)", flat["volume"]),
        ("Volatility (ATR %)", flat["atr_pct"]),
        ("RSI divergence", flat["divergence"]),
        ("Nearest floors", flat["supports"]),
        ("Nearest ceilings", flat["resistances"]),
        ("52-week range", flat["week52_low"] + "  to  " + flat["week52_high"]),
    ], styles, avail_w, ncols=2))
    story.append(Spacer(1, 5))

    story.append(Paragraph("The trade plan", styles["h3"]))
    plan = Table([[_plan_table(flat, styles, avail_w * 0.50),
                   Paragraph("<b>Outlook:</b> " + _ascii(t.get("outlook_label"))
                             + "<br/><br/>" + _ascii(t.get("outlook_blurb"))
                             + "<br/><br/><b>Entry note:</b> "
                             + _ascii(t.get("entry_note")), styles["body"])]],
                 colWidths=[avail_w * 0.52, avail_w * 0.48])
    plan.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                              ("LEFTPADDING", (1, 0), (1, 0), 10)]))
    story.append(plan)

    story.append(Spacer(1, 5))
    pc = Table([[[Paragraph("What's going right", styles["h3"]),
                  _bullets(t.get("pros"), styles, "Nothing strongly positive.")],
                 [Paragraph("What to watch", styles["h3"]),
                  _bullets(t.get("cons"), styles, "No major warning flags.")]]],
               colWidths=[avail_w / 2, avail_w / 2])
    pc.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 0),
                            ("RIGHTPADDING", (0, 0), (0, 0), 10)]))
    story.append(pc)

    if _s(row.get("psx_url")) != DASH:
        story += [Spacer(1, 4),
                  Paragraph("Verify every figure at: " + _ascii(row.get("psx_url")),
                            styles["note"])]
    return story


def build_pdf(rows: List[Dict], meta: Dict = None) -> bytes:
    if not HAVE_PDF:
        raise RuntimeError(
            "PDF export needs the 'reportlab' package. Close the app, run:\n"
            "    pip install reportlab\nand launch again.")
    meta = meta or {}
    rows = rows or []
    st = _pdf_styles()
    buf = io.BytesIO()
    doc = _Doc(buf, meta)
    land_w = landscape(A4)[0] - 24 * mm
    port_w = A4[0] - 28 * mm

    stamp = meta.get("as_of") or _dt.date.today().isoformat()
    mode = meta.get("mode") or ""
    story: List[Any] = []

    # ---- cover / summary (landscape) -------------------------------------
    story.append(Paragraph("PSX.SCORE - Stock Analysis Report", st["h1"]))
    story.append(Paragraph(
        _ascii(f"Fundamental &amp; Technical X-ray  |  {len(rows)} "
               f"compan{'y' if len(rows) == 1 else 'ies'}  |  figures describe "
               f"the close of {stamp}"
               + (f"  |  {mode} data" if mode else "")
               + f"  |  generated {_dt.datetime.now().strftime('%d %b %Y, %H:%M')}"),
        st["sub"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph("At a glance", st["h2"]))
    story.append(_summary_table(rows, st, land_w))
    story.append(Spacer(1, 7))
    story.append(Paragraph(
        "Every figure above is the one shown on screen - nothing is estimated, "
        "recalculated or filled in. Blank readings appear as a dash. Buy zones, "
        "stop-loss and targets are chart levels for planning only and are not a "
        "recommendation to trade.", st["note"]))

    # ---- per-stock detail (portrait) -------------------------------------
    if rows:
        story.append(NextPageTemplate("port"))
        story.append(PageBreak())
        for i, row in enumerate(rows):
            story += _stock_section(row, st, port_w)
            if i < len(rows) - 1:
                story.append(PageBreak())

    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
def build(fmt: str, rows: List[Dict], meta: Dict = None) -> tuple:
    """→ (bytes, mime, filename)"""
    fmt = (fmt or "").lower().strip()
    meta = meta or {}
    rows = rows or []
    base = meta.get("filename") or "PSX-SCORE"
    if len(rows) == 1 and rows[0].get("symbol"):
        base = f"PSX-SCORE_{rows[0]['symbol']}"
    stamp = _dt.date.today().strftime("%Y-%m-%d")

    if fmt in ("xlsx", "excel", "xls"):
        return (build_workbook(rows, meta),
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet",
                f"{base}_{stamp}.xlsx")
    if fmt == "pdf":
        return (build_pdf(rows, meta), "application/pdf",
                f"{base}_{stamp}.pdf")
    raise ValueError("format must be 'pdf' or 'xlsx'")
